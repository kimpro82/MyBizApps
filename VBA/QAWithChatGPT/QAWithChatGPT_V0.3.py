"""
excel_chat_gpt.py / 2023.08.12

openpyxl, requests, pyyaml 라이브러리를 이용하여 ExcelChatGPT 0.3 버전의 기능을 구현한 파이썬 코드입니다.
엑셀 파일에서 필요한 데이터를 불러와 ChatGPT API에 요청을 보내고, 그 결과를 YAML 파일로 저장합니다.

주의: 코드 실행 전 필요한 라이브러리를 설치해야 합니다. (pip install openpyxl requests pyyaml)
"""

import pprint
import openpyxl
import requests
import yaml

# 경로 설정
FILE_PATH       = "./QAWithChatGPT.xlsm"
SHEET_NAME      = "ChatGPT0.3"
SAVE_FILE_PATH  = "./QAWithChatGPT_V0.3.yml"

# 엑셀 파일 불러오기
wb          = openpyxl.load_workbook(FILE_PATH)
ws          = wb[SHEET_NAME]

# 엑셀에서 데이터 불러오기
model       = ws["C5"].value
tokens      = ws["E5"].value
endpoint    = ws["F7"].value
api_key     = ws["C8"].value
question    = ws["C11"].value

# Json 요청 데이터 생성
request_data = {
    "model"     : model,
    "messages"  : [
        {
            "role"      : "system",
            "content"   : "You are a helpful assistant."
        },
        {
            "role"      : "user",
            "content"   : question
        }
    ],
    "max_tokens": tokens,
    "n"         : 1
}

headers = {
    "Content-Type"  : "application/json",
    "Authorization" : f"Bearer {api_key}"
}

# ChatGPT API 요청
url         = f"https://api.openai.com{endpoint}"
response    = requests.post(url, json=request_data, headers=headers, timeout=5)
answer      = response.json()
# content   = answer['choices'][0]['message']['content']
# print("content(raw)         :", content)

# YAML 파일로 저장
with open(SAVE_FILE_PATH, "w", encoding='utf-8') as file:
    yaml.dump(answer, file)

# 테스트
if __name__ == '__main__':
    pprint.pprint(request_data)
    print()
    pprint.pprint(response.json())
    print()
